#!/usr/bin/env python3
"""福岡市の財政データを「地方財政状況資料集」Excel から取得し、表示用JSONを生成する。

データ源: 福岡市公式「地方財政状況資料集（各年度決算）」Excel
  例 令和6年度: https://www.city.fukuoka.lg.jp/zaisei/zaisei/shisei/documents/6zaiseijoukyoushiryoushuu.xlsx
  （Nzaiseijoukyoushiryoushuu.xlsx の N = 令和の年。各年度決算ページからもリンクされる）

抽出方針（読み違い防止）:
  - 歳出目的別 / 歳入財源別 は「普通会計の状況」シートを **ラベルベース** で抽出
    （行がずれても項目名で特定。総額・人口・指標は「総括表」から）
  - **目的別歳出の合計が歳出合計と一致すること** を検証してから出力（不一致なら中断）
  - 金額は千円。歳入の各種交付金（小項目群）は「各種交付金等」に集約（依存財源に分類される）

実行: python3 scripts/fetch-fukuoka-finance.py   （要 openpyxl / ネットワーク）
出力: web/src/features/city-finance/data/fukuoka-finance.json

詳細は .claude/skills/fukuoka-finance-map/SKILL.md を参照。
"""
from __future__ import annotations

import datetime
import io
import json
import os
import re
import unicodedata
import urllib.request

import openpyxl

UA = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120 Safari/537.36"
DOC_BASE = "https://www.city.fukuoka.lg.jp/zaisei/zaisei/shisei/documents"
# 高齢化率: 福岡市オープンデータ(BODIK) 住民基本台帳・年齢別人口（合算・年次一括）
BODIK_API = "https://data.bodik.jp/api/3/action"
AGING_DATASET = "401307_fukuokacityjukijinkou"
OUT_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "web/src/features/city-finance/data/fukuoka-finance.json",
)

# 令和の年 -> 西暦年度
YEARS = {6: 2024, 5: 2023, 4: 2022, 3: 2021, 2: 2020}

# 目的別歳出（13区分・「普通会計の状況」の目的別ブロック）
PURPOSE = [
    "議会費", "総務費", "民生費", "衛生費", "労働費", "農林水産業費", "商工費",
    "土木費", "消防費", "教育費", "災害復旧費", "公債費", "諸支出金",
]
# 歳入の主要財源（款）。残りは「各種交付金等」に集約
REVENUE = [
    "地方税", "地方譲与税", "地方交付税", "分担金・負担金", "使用料", "手数料",
    "国庫支出金", "都道府県支出金", "財産収入", "寄附金", "繰入金", "繰越金",
    "諸収入", "地方債",
]
INDICATORS = ["経常収支比率", "財政力指数", "実質公債費比率", "将来負担比率"]


def get_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    return urllib.request.urlopen(req, timeout=60).read()


def first_num_right(ws, r: int, c: int):
    for cc in range(c + 1, ws.max_column + 1):
        v = ws.cell(r, cc).value
        if isinstance(v, (int, float)):
            return v
    return None


def find_label_value(ws, label: str):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().replace("\n", "") == label:
                n = first_num_right(ws, r, c)
                if n is not None:
                    return n
    return None


def first_occurrence_map(ws, labels: set[str]) -> dict[str, float]:
    found: dict[str, float] = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                t = v.strip().replace("\n", "")
                if t in labels and t not in found:
                    n = first_num_right(ws, r, c)
                    if n is not None:
                        found[t] = n
    return found


def latest_juki_population(sk) -> float | None:
    """総括表から住民基本台帳人口（最新の1月1日時点）を取得。"""
    best_year, best_val = -1, None
    for r in range(1, sk.max_row + 1):
        for c in range(1, sk.max_column + 1):
            v = sk.cell(r, c).value
            if isinstance(v, str):
                m = re.search(r"令0?(\d+)\.01\.01", v.replace(" ", ""))
                if m:
                    val = first_num_right(sk, r, c)
                    if val and int(m.group(1)) > best_year:
                        best_year, best_val = int(m.group(1)), val
    return best_val


def fetch_aging_rate() -> list[dict]:
    """福岡市オープンデータ(BODIK)の年齢別人口（合算）から各年の65歳以上比率(%)を算出。
    各年ファイルの最初の月末シート・福岡市総数(C列)を用いる。合算ファイルがある年のみ。
    """
    pkg = json.loads(
        get_bytes(f"{BODIK_API}/package_show?id={AGING_DATASET}").decode()
    )
    five = re.compile(r"^(65|70|75|80|85|90|95)~")

    def norm(s: object) -> str:
        return unicodedata.normalize("NFKC", str(s)).replace(" ", "")

    out: list[dict] = []
    for year in sorted(YEARS.values()):
        url = next(
            (
                r["url"]
                for r in pkg["result"]["resources"]
                if f"{year}nenreibetu-sousuu" in r["url"]
            ),
            None,
        )
        if url is None:
            continue  # 合算ファイル未公開の年はスキップ
        wb = openpyxl.load_workbook(io.BytesIO(get_bytes(url)), data_only=True)
        ws = wb[wb.sheetnames[0]]
        total: float | None = None
        senior = 0.0
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if not isinstance(a, str):
                continue
            label = norm(a)
            value = ws.cell(r, 3).value  # C列=福岡市総数
            if not isinstance(value, (int, float)):
                continue
            if "総数" in label and total is None:
                total = value
            if five.match(label) or (label.startswith("100") and "以上" in label):
                senior += value
        if total:
            out.append({"year": year, "value": round(senior / total * 100, 1)})
            print(f"  高齢化率 {year}: {out[-1]['value']}%")
    return out


def main() -> None:
    gen = {"歳入総額": [], "歳出総額": []}
    rev = {k: [] for k in REVENUE}
    other: list[dict] = []
    exp = {k: [] for k in PURPOSE}
    pop: list[dict] = []
    indicators = {k: [] for k in INDICATORS}

    for n in sorted(YEARS, reverse=True):
        fy = YEARS[n]
        wb = openpyxl.load_workbook(
            __import__("io").BytesIO(get_bytes(f"{DOC_BASE}/{n}zaiseijoukyoushiryoushuu.xlsx")),
            data_only=True,
        )
        ws = wb["普通会計の状況"]
        sk = wb["総括表"]

        revenue_total = find_label_value(sk, "歳入総額")
        expense_total = find_label_value(sk, "歳出総額")
        if revenue_total is None or expense_total is None:
            raise SystemExit(f"R{n}: 総括表から総額を取得できません")

        # 目的別歳出 + 合計突合検証
        ef = first_occurrence_map(ws, set(PURPOSE))
        purpose_sum = sum(ef.values())
        if abs(purpose_sum - expense_total) >= 1000:
            raise SystemExit(
                f"R{n}: 目的別歳出の合計({purpose_sum:,})が歳出合計({expense_total:,})と不一致"
            )

        rf = first_occurrence_map(ws, set(REVENUE))
        # 抽出ドリフト検知: 主要費目/財源が欠けたら中断（黙って各種交付金等に吸収させない）
        missing_p = [k for k in PURPOSE if k not in ef]
        missing_r = [k for k in REVENUE if k not in rf]
        if missing_p or missing_r:
            raise SystemExit(
                f"R{n}: 抽出できないラベル 目的別={missing_p} 歳入={missing_r}"
            )
        residual = revenue_total - sum(rf.values())
        if not 0 <= residual <= revenue_total * 0.15:
            raise SystemExit(
                f"R{n}: 各種交付金等の残差({residual:,})が想定範囲外（歳入総額の0〜15%）"
            )

        gen["歳入総額"].append({"year": fy, "value": revenue_total})
        gen["歳出総額"].append({"year": fy, "value": expense_total})
        for k in PURPOSE:
            if k in ef:
                exp[k].append({"year": fy, "value": ef[k]})
        for k in REVENUE:
            if k in rf:
                rev[k].append({"year": fy, "value": rf[k]})
        other.append({"year": fy, "value": residual})
        pop_val = latest_juki_population(sk)
        if pop_val is None:
            raise SystemExit(f"R{n}: 総括表から住基人口を取得できません")
        pop.append({"year": fy, "value": pop_val})
        for key in INDICATORS:
            indicators[key].append({"year": fy, "value": find_label_value(sk, key)})
        print(f"R{n}/{fy}: 検証OK 歳入{revenue_total:,} 歳出{expense_total:,}")

    def sort_series(d):
        return [
            {"item": k, "values": sorted(v, key=lambda x: x["year"])}
            for k, v in d.items()
        ]

    data = {
        "source": {
            "name": "福岡市オープンデータ（地方財政状況資料集）",
            "url": "https://www.city.fukuoka.lg.jp/zaisei/zaisei/shisei/R6aramashi.html",
            "datasetId": "fukuoka-zaiseijoukyoushiryoushuu",
            "fetchedAt": datetime.date.today().isoformat(),
        },
        "unit": "thousand_yen",
        "years": sorted(YEARS.values()),
        "generalAccount": sort_series(gen),
        "revenue": sort_series(rev)
        + [{"item": "各種交付金等", "values": sorted(other, key=lambda x: x["year"])}],
        "expenditure": sort_series(exp),
        "population": sorted(pop, key=lambda x: x["year"]),
        # 財政指標（%・指数）。総括表より。
        "indicators": [
            {
                "item": k,
                "values": sorted(
                    [e for e in v if e["value"] is not None],
                    key=lambda x: x["year"],
                ),
            }
            for k, v in indicators.items()
        ],
        # 高齢化率（65歳以上比率・%）。福岡市オープンデータ 年齢別人口より。
        "agingRate": fetch_aging_rate(),
    }
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
